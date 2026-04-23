import json
import pathlib
from openpyxl import load_workbook

base = pathlib.Path(r"c:/Users/Khaled/Documents/GitHub/ABET/Docs")
files = [
    "AIU_STUDENT_GRADES_11082025.xlsx",
    "AIU_ENROLLMENT_2242.xlsx",
    "SO-Coverage-V2 - Copy.xlsx",
]
out = {}
for fn in files:
    p = base / fn
    wb = load_workbook(p, data_only=True)
    meta = {}
    for ws in wb.worksheets:
        rows = []
        for r in ws.iter_rows(min_row=1, max_row=8, values_only=True):
            vals = [None if v is None else str(v) for v in r]
            if any(v is not None for v in vals):
                rows.append(vals)
        meta[ws.title] = rows
    out[fn] = meta

print(json.dumps(out, indent=2, ensure_ascii=True))
